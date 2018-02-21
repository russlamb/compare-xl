import pyodbc
import datetime
import openpyxl


class SqlToXl():
    """Use to connect to database, run sql, write results to Excel file"""

    def __init__(self, connection_string):
        self.connection_string = connection_string

    def save_sql(self, sql, filename, sheetname="Sheet1"):
        cnxn = pyodbc.connect(self.connection_string)
        cursor = cnxn.cursor()
        cursor.execute(sql)
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = sheetname

        columns = [column[0] for column in cursor.description]

        rowid = 1
        colid = 1
        for col in columns:
            ws.cell(row=rowid, column=colid).value = col
            colid += 1

        for row in cursor:
            rowid += 1
            colid = 1
            for col in row:
                ws.cell(row=rowid, column=colid).value = col
                colid += 1

        wb.save(filename)
        print("Saved {} @ {}".format(filename,datetime.datetime.now() ))