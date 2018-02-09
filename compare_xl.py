import itertools

import openpyxl
import datetime


class CompareXl():
    """Compare two excel files and save a new file with the differences"""

    def __init__(self, file1, file2, sheetname1, sheetname2):
        self.file1 = file1
        self.file2 = file2
        self.wb1 = openpyxl.load_workbook(file1)
        self.wb2 = openpyxl.load_workbook(file2)
        self.sheet1 = self.wb1[sheetname1]
        self.sheet2 = self.wb2[sheetname2]

    def compare_index_and_save(self, index, filename):
        """Get differences using index and save to file"""
        diffs = self.compare_index(index)
        self.save_differences(filename, diffs)

    def compare_and_save(self, filename):
        """Get differences and save to file"""
        diffs = self.compare()
        self.save_differences(filename, diffs)

    def compare_index(self, index_column):
        """compare excel files using an index to align rows.  index column is one based"""
        print("Compare using index {} @ {}".format(index_column,datetime.datetime.now()))
        differences = []

        index_dict1 = dict()
        index_dict2 = dict()

        index_set1 = set()
        index_set2 = set()

        index_letter = openpyxl.utils.get_column_letter(index_column)
        for cell in self.sheet1[index_letter]:
            index_dict1[cell.value] = cell.row
            index_set1.add(cell.value)
        for cell in self.sheet2[index_letter]:
            index_dict2[cell.value] = cell.row
            index_set2.add(cell.value)

        index_intersect = index_set1.intersection(index_set2)

        index_list = list(index_intersect)

        # info for user regarding index
        print("# index values not in both files={}".format(len(index_set1.symmetric_difference(index_set2))))
        print("# index values in both files={}".format(len(index_list)))
        print("# index values on different rows={}".format(
            len([i for i in index_list if index_dict1[i] != index_dict2[i]])))

        for i in index_list:

            index_val1 = index_dict1[i]
            index_val2 = index_dict2[i]

            for col in range(1, self.sheet1.max_column):
                col1 = self.sheet1.cell(row=index_val1, column=col)
                col2 = self.sheet2.cell(row=index_val2, column=col)
                if not col1.value == col2.value:
                    column_header = self.sheet1[col1.column + "1"].value
                    differences.append([column_header, col1.row, col1.column, col1.value, col2.value])

        return differences

    def compare(self):
        """Compare XLSX cells between two files.  Only compare cells that are contained in both."""
        print("Compare direct (no index) @ {}".format( datetime.datetime.now()))
        differences = []
        for (row1, row2) in zip(self.sheet1.iter_rows(), self.sheet2.iter_rows()):
            for (col1, col2) in zip(row1, row2):
                if not col1.value == col2.value:
                    column_header = self.sheet1[col1.column + "1"].value
                    differences.append([column_header, col1.row, col1.column, col1.value, col2.value])

        return differences

    def compare_uneven(self):
        """compare all cells of XLSX files, even if one is longer than the other"""
        print("Compare uneven lengths @ {}".format(datetime.datetime.now()))
        for (row1, row2) in itertools.zip_longest(self.sheet1.iter_rows(), self.sheet2.iter_rows()):
            for (col1, col2) in itertools.zip_longest(row1, row2):

                if not col1.value == col2.value:
                    print("{},{}  == '{}' <> '{}'".format(col1.row, col1.column, col1.value, col2.value))

    def save_differences(self, output_file, differences):
        """Save list of lists to an XLSX file"""
        print("Save differences to {} @ {}".format(output_file,datetime.datetime.now()))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "differences"

        rowid = 1
        colid = 1

        header = ["header", "row", 'column', self.file1, self.file2]

        for col in header:
            ws.cell(row=rowid, column=colid).value = col
            colid += 1

        rowid += 1
        for row in differences:
            colid = 1
            for item in row:
                ws.cell(row=rowid, column=colid).value = item
                colid += 1
            rowid += 1

        wb.save(output_file)
